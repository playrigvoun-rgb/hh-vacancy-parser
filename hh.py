from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

options = Options()
options.add_argument('--disable-blink-feauters=AutomationControlled')
# options - настройки для нашего браузера.add_argument - добавляем настройку: --disable-blink-feauters - указываем список функций(feauters)для блинк(движок хрома)=меняем параметр, чтобы сайт не думал что мы бот
options.add_experimental_option('excludeSwitches', ['enable-automation'])
# добавляем экспериментальную опцию для браузера, switches - параметры ком. строки при запуске передаются браузеру. exclude - просим исключить какие то параметры, enable-automation - просим, чтобы браузер исключил параметр, что мы зашли с бота в браузер, то есть чтоб он ничего не урезал
options.add_experimental_option('useAutomationExtension', False)
# отрубаем расширение для хрома, которое запускается по умолчанию с селениумом, это расширение настакже выдает


print('----------------------------------------------------------------------')
search_vacancy = input('Введите название профессии: ')
print('====------------------------------------------------------------------')
time.sleep(0.3)
print('===============-------------------------------------------------------')
time.sleep(0.3)
print('===============================---------------------------------------')
time.sleep(0.3)
print('===========================================---------------------------')
time.sleep(0.3)
print('========================================================--------------')
time.sleep(0.3)
print('================================================================------')
time.sleep(0.3)
print('======================================================================')
print(f'{search_vacancy} - профессия выбрана.')
print('----------------------------------------------------------------------')
region_input = input('Введите регион поиска: ')
print('====------------------------------------------------------------------')
time.sleep(0.3)
print('===============-------------------------------------------------------')
time.sleep(0.3)
print('===============================---------------------------------------')
time.sleep(0.3)
print('===========================================---------------------------')
time.sleep(0.3)
print('========================================================--------------')
time.sleep(0.3)
print('================================================================------')
time.sleep(0.3)
print('======================================================================')
print(f'{region_input} - регион выбран.')
print('----------------------------------------------------------------------')
number_of_pages_f = input('Введите кол-во страниц, которое нужно спарсить(примените фильтры выше на сайте https://hh.ru и посмотрите сколько страниц есть по этим фильтрам. Если хотите спарсить все страницы - введите номер последней страницы): ')
print('====------------------------------------------------------------------')
time.sleep(0.3)
print('===============-------------------------------------------------------')
time.sleep(0.3)
print('===============================---------------------------------------')
time.sleep(0.3)
print('===========================================---------------------------')
time.sleep(0.3)
print('========================================================--------------')
time.sleep(0.3)
print('================================================================------')
time.sleep(0.3)
print('======================================================================')
print('Все настройки применены, направляю вас на сайт...')
driver = webdriver.Chrome(options=options)

driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
# обращаемся к драйверу который и управляет браузером, ее запускает. execute_script - говорим драйверу выполнить этот код на этой странице, далее - javascript: Object - глобальный обьект в джава скрипт, в нем есть методы для работы с обьектами. defineProperty - метод определяет новое свойство у обьекта или меняет существующее. navigator - спец обьект, содержит инфу о браузере и операционной системе. 'webdriver' - свойство внутри navigator - он появляется при запуске браузера от селениум, он равен true при запуске через селениум и потому палит нас, {get: () => undefined} то есть мы делаем значение webdriver - undefined а не true что нас скрывает

driver.get('https://hh.ru/')



all_jobs = []

#time.sleep(2)

title = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.ID, 'a11y-search-input'))
)
title.click()

title.send_keys(f'{search_vacancy}')
#time.sleep(1)
title.send_keys(Keys.RETURN)


number_of_pages = int(number_of_pages_f)
number_of_pages_n = number_of_pages + 1

cross = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="signup-modal-close"]'))
)
cross.click()

time.sleep(2)
print(driver.current_url)
#time.sleep(1.5)



# тем самым получили текущий юрл
page = 0
a = 1




while a != number_of_pages_n:
    if page == 0:
        page = str(page)
        real_url_f = str(driver.current_url)
        real_url = real_url_f
        real_url = real_url.replace('velikie-luki.', '')
        page = int(page)
        print(real_url)
        time.sleep(1)
        print(f'{page+1} страница')
        driver.get(real_url)
        time.sleep(2)
        region_tag = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="search-filter-area-chip"]'))
        )

        region = region_tag.text.strip()
        if region != region_input:
            
            
        
            region_tag.click()
            time.sleep(0.5)

            del_city = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="cell-text"]'))
            )
            del_city.click()
            search = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="search-filter-area-input"]'))
            )
            search.send_keys(f'{region_input}')
            time.sleep(0.5)

            your_region = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="cell-text"]'))
            )
            your_region.click()
            save = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="search-filter-area-apply-button"]'))
            )
            save.click()
       

        time.sleep(2)
   
    
        
        vacancies = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, 'vacancy-card--n77Dj8TY8VIUF0yM'))
        )
    

        for vacancy in vacancies:
            try:
                title = vacancy.find_element(By.CSS_SELECTOR, '[data-qa="serp-item__title-text"]')
            except:
                title = 'Не найдено'
            try:
                experience = vacancy.find_element(By.CLASS_NAME, 'magritte-tag__label___YHV-o_5-3-14')
            except:
                experience = 'Нет информации'


            try:
                salary_voc = vacancy.find_elements(By.CLASS_NAME, 'magritte-text___pbpft_5-3-7')
                salary = salary_voc[2].text.strip()
                if "месяц" in salary:
                    salary = salary
                else:
                    salary = 'Уровень дохода не указан'

            except:
                salary = 'Уровень дохода не указан'

            if 'на руки' in salary:
                salary = salary.replace(', на руки', '')
                salary_category = 'На руки'
            elif 'до вычета налогов' in salary:
                salary = salary.replace(', до вычета налогов', '')
                salary_category = 'До вычета налогов'
            else:
                salary = salary
                salary_category = 'Не указан'

            company = vacancy.find_element(By.CSS_SELECTOR, '[data-qa="vacancy-serp__vacancy-employer-text"]').text

            address = vacancy.find_element(By.CSS_SELECTOR, '[data-qa="vacancy-serp__vacancy-address"]').text

            link = vacancy.find_element(By.CLASS_NAME, 'magritte-link___b4rEM_7-1-33').get_attribute('href').replace('velikie-luki.', '')

            all_jobs.append({
                'Вакансия' : title.text,
                'Компания' : company,
                'Зарплатная вилка' : salary,
                'Тип выплаты' : salary_category,
                'Опыт' : experience.text,
                'Адрес компании' : address,
                'Ссылка на вакансию' : link
            })


            print(f'{title.text} - {company} - {salary} - {salary_category} - {experience.text} - {address} - {link}')
        a += 1

        page += 1

    
    else:
        page = str(page)
        real_url_f = str(driver.current_url)
        real_url = real_url_f + '&page=' + page
        real_url = real_url.replace('velikie-luki.', '')
        page = int(page)
        print(real_url)
        time.sleep(1)
        print(f'{page+1} страница')
        driver.get(real_url)
        region_tag = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="search-filter-area-chip"]'))
        )

        time.sleep(2)

        region = region_tag.text.strip()
        if region != region_input:
            
            
        
            region_tag.click()
            time.sleep(0.5)

            del_city = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="cell-text"]'))
            )
            del_city.click()
            search = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="search-filter-area-input"]'))
            )
            search.send_keys(f'{region_input}')
            time.sleep(0.5)

            your_region = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="cell-text"]'))
            )
            your_region.click()
            save = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="search-filter-area-apply-button"]'))
            )
            save.click()
       


   
    
        
        vacancies = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, 'vacancy-card--n77Dj8TY8VIUF0yM'))
        )
    

        for vacancy in vacancies:
            try:
                title = vacancy.find_element(By.CSS_SELECTOR, '[data-qa="serp-item__title-text"]')
            except:
                title = 'Не найдено'
            try:
                experience = vacancy.find_element(By.CLASS_NAME, 'magritte-tag__label___YHV-o_5-3-14')
            except:
                experience = 'Нет информации'


            try:
                salary_voc = vacancy.find_elements(By.CLASS_NAME, 'magritte-text___pbpft_5-3-7')
                salary = salary_voc[2].text.strip()
                if "месяц" in salary:
                    salary = salary
                else:
                    salary = 'Уровень дохода не указан'

            except:
                salary = 'Уровень дохода не указан'

            if 'на руки' in salary:
                salary = salary.replace(', на руки', '')
                salary_category = 'На руки'
            elif 'до вычета налогов' in salary:
                salary = salary.replace(', до вычета налогов', '')
                salary_category = 'До вычета налогов'
            else:
                salary = salary
                salary_category = 'Не указан'

            company = vacancy.find_element(By.CSS_SELECTOR, '[data-qa="vacancy-serp__vacancy-employer-text"]').text

            address = vacancy.find_element(By.CSS_SELECTOR, '[data-qa="vacancy-serp__vacancy-address"]').text

            link = vacancy.find_element(By.CLASS_NAME, 'magritte-link___b4rEM_7-1-33').get_attribute('href').replace('velikie-luki.', '')

            all_jobs.append({
                'Вакансия' : title.text,
                'Компания' : company,
                'Зарплатная вилка' : salary,
                'Тип выплаты' : salary_category,
                'Опыт' : experience.text,
                'Адрес компании' : address,
                'Ссылка на вакансию' : link
            })


            print(f'{title.text} - {company} - {salary} - {salary_category} - {experience.text} - {address} - {link}')
        a += 1

        page += 1


driver.quit()


df = pd.DataFrame(all_jobs, columns=['Вакансия', 'Компания',  'Зарплатная вилка', 'Тип выплаты', 'Опыт', 'Адрес компании', 'Ссылка на вакансию'])
df.to_excel('jobs_hh_python.xlsx',index=False)

wb = load_workbook('jobs_hh_python.xlsx')
ws = wb.active



# --- 1. ПРИМЕНЯЕМ ДИЗАЙН К ЗАГОЛОВКАМ (шапка) ---
# Цвета: тёмно-синий фон, белый жирный текст, центр
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_alignment = Alignment(horizontal='center', vertical='center')

for cell in ws[1]:  # Первая строка — заголовки
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# --- 2. ЧЕРЕДОВАНИЕ СТРОК (зебра) ---
# Белый и светло-серый для читаемости
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        if cell.row % 2 == 0:  # Чётные строки
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        # Нечётные остаются белыми

# --- 3. ВЫРАВНИВАНИЕ ДАННЫХ ---
# Названия — по левому краю, Цена и Рейтинг — по центру
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        # Если столбец A (Название) — влево
        if cell.column == 1:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        # Если столбец B (Цена) или C (Рейтинг) — по центру
        elif cell.column in [2, 3]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Остальные (Наличие, Ссылка) — по центру
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# --- 4. ГРАНИЦЫ ---
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        cell.border = thin_border

# --- 5. АВТОШИРИНА СТОЛБЦОВ ---
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# --- 6. ВКЛЮЧАЕМ ФИЛЬТРЫ ---
ws.auto_filter.ref = ws.dimensions

# --- 7. ЗАКРЕПЛЯЕМ ШАПКУ (чтобы видно было при скролле) ---
ws.freeze_panes = 'A2'



for row in range(2, ws.max_row + 1):
    cell = ws.cell(row, 7)
    if cell.value and cell.value.startswith('http'):
        cell.hyperlink = cell.value
        cell.font = Font(color='0000FF', underline='single')

color_map = {
    '⭑⭑⭑⭑⭑' : '00B050',
    '⭑⭑⭑⭑⭒' : '92D050',
    '⭑⭑⭑⭒⭒' : 'FFC000',
    '⭑⭑⭒⭒⭒' : 'ED7D31',
    '⭑⭒⭒⭒⭒' : 'FF0000'
}

rating_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(1, col).value == 'Рейтинг':
        rating_col = col
        break
if rating_col:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, rating_col)
        rating_value = cell.value
        if rating_value in color_map:
            cell.fill = PatternFill(start_color=color_map[rating_value], end_color=color_map[rating_value], fill_type = 'solid')

wb.save('jobs_hh_python.xlsx')

print(f'Готово! Сохранено {len(all_jobs)} в этот файл')